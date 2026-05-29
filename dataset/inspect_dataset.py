import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    path = Path(__file__).resolve().parent / "B2B_Sales_Data.xlsx"
    out_path = Path(__file__).resolve().parent / "inspect_dataset_output.json"

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    columns = [str(c).strip() if c is not None else "" for c in header]

    null_counts = {c: 0 for c in columns}
    type_sets = {c: set() for c in columns}
    uniques = {c: set() for c in columns}
    min_values = {c: None for c in columns}
    max_values = {c: None for c in columns}

    counter_cols = {
        "seniority_level",
        "company_industry",
        "company_size_category",
        "lead_source",
        "webinar_attendance",
        "meeting_scheduled",
        "follow_up_status",
        "estimated_budget",
        "purchase_timeline",
        "country",
        "department",
        "job_title",
        "company_size_range",
        "converted",
    }
    counters = {c: Counter() for c in columns if c in counter_cols}

    row_count = 0
    for row in rows_iter:
        row_count += 1
        for idx, col in enumerate(columns):
            val = row[idx] if idx < len(row) else None
            if val is None or val == "":
                null_counts[col] += 1
                continue
            type_sets[col].add(type(val).__name__)
            try:
                uniques[col].add(val)
            except TypeError:
                uniques[col].add(str(val))
            if col in counters:
                counters[col][str(val)] += 1
            if isinstance(val, (int, float)) or type(val).__name__ in {"datetime", "date"}:
                if min_values[col] is None or val < min_values[col]:
                    min_values[col] = val
                if max_values[col] is None or val > max_values[col]:
                    max_values[col] = val

    def infer_dtype(type_names: set[str]) -> str:
        if not type_names:
            return "unknown"
        if type_names == {"bool"}:
            return "bool"
        if "datetime" in type_names or "date" in type_names:
            return "datetime"
        if type_names.issubset({"int", "float"}):
            return "float" if "float" in type_names else "int"
        if type_names.issubset({"int"}):
            return "int"
        if type_names.issubset({"float"}):
            return "float"
        return "string"

    dtypes = {c: infer_dtype(type_sets[c]) for c in columns}

    summary = {
        "file": str(path),
        "sheet": wb.sheetnames[0],
        "shape": [int(row_count), int(len(columns))],
        "columns": columns,
        "dtypes": dtypes,
        "python_type_sets": {c: sorted(type_sets[c]) for c in columns},
        "null_counts": {c: int(null_counts[c]) for c in columns},
        "nunique": {c: int(len(uniques[c])) for c in columns},
        "min_max": {
            c: {"min": str(min_values[c]) if min_values[c] is not None else None, "max": str(max_values[c]) if max_values[c] is not None else None}
            for c in columns
            if min_values[c] is not None or max_values[c] is not None
        },
    }

    if "converted" in columns:
        vc = counters.get("converted", Counter())
        summary["target"] = {
            "column": "converted",
            "value_counts": {str(k): int(v) for k, v in vc.items()},
        }

    categorical_priority = [
        "seniority_level",
        "company_industry",
        "company_size_category",
        "lead_source",
        "webinar_attendance",
        "meeting_scheduled",
        "follow_up_status",
        "estimated_budget",
        "purchase_timeline",
        "country",
        "department",
        "job_title",
        "company_size_range",
    ]

    categorical_uniques = {}
    for c in categorical_priority:
        if c not in columns:
            continue
        values_sorted = sorted({str(v) for v in uniques[c]})
        categorical_uniques[c] = {
            "nunique": int(len(uniques[c])),
            "values_preview": values_sorted[:50],
        }

    inferred_categorical = []
    for c in columns:
        if dtypes[c] in {"string", "bool"}:
            inferred_categorical.append(c)
            continue
        if len(uniques[c]) <= 15:
            inferred_categorical.append(c)

    summary["categorical"] = {
        "priority_columns": categorical_uniques,
        "inferred_categorical_columns": inferred_categorical,
        "counters_top10": {
            c: counters[c].most_common(10)
            for c in counters
        },
    }

    wb.close()

    out_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(str(out_path))


if __name__ == "__main__":
    main()

